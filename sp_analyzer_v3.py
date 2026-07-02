#!/usr/bin/env python3
"""
Stored Procedure Analyzer v3 — Universal SQL Dialect
- Physical tables only in main analysis
- Temp tables (#) shown as warnings/notes only
- New Schema Breakdown tab: Schema → Tables → Columns
- Hybrid deterministic regex engine (no LLM)

Usage:
    python sp_analyzer_v3.py                         # paste SP(s) interactively
    python sp_analyzer_v3.py my_sp.sql               # single file
    python sp_analyzer_v3.py sp1.sql sp2.sql ...     # multiple files
    python sp_analyzer_v3.py --dialect tsql my.sql   # force dialect
    python sp_analyzer_v3.py --output report.xlsx my.sql
"""

import re
import sys
import argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Palette ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F3864"
C_HEADER_FG  = "FFFFFF"
C_SUBHDR_BG  = "2E75B6"
C_SUBHDR_FG  = "FFFFFF"
C_ALT_ROW    = "DCE6F1"
C_WHITE      = "FFFFFF"
C_WARN_BG    = "FFF2CC"
C_WARN_FG    = "7F6000"
C_SCHEMA_BG  = "E2EFDA"   # green tint for schema grouping rows
C_SCHEMA_FG  = "375623"

OP_COLORS = {
    "SELECT":   "375623",
    "INSERT":   "833C00",
    "UPDATE":   "7F6000",
    "DELETE":   "C00000",
    "MERGE":    "3A3268",
    "TRUNCATE": "420000",
    "DROP":     "595959",
    "CREATE":   "17375E",
    "CTE":      "17375E",
}


# ── Styling ───────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()
    return cell

def style_data(cell, row_idx):
    bg = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def style_warn(cell):
    cell.fill      = PatternFill("solid", fgColor=C_WARN_BG)
    cell.font      = Font(bold=True, color=C_WARN_FG, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def style_schema_group(cell):
    cell.fill      = PatternFill("solid", fgColor=C_SCHEMA_BG)
    cell.font      = Font(bold=True, color=C_SCHEMA_FG, name="Arial", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = thin_border()

def op_color(ops):
    for op in ["DELETE","TRUNCATE","DROP","MERGE","UPDATE","INSERT","SELECT","CREATE","CTE"]:
        if op in ops:
            return OP_COLORS.get(op, "404040")
    return "404040"

def strip_name(n):
    return re.sub(r'[\[\]"`]', '', str(n)).strip()

def is_temp(name):
    return str(name).lstrip('[').startswith('#')

def is_var(name):
    return str(name).startswith('@')


# ── Dialect Detection ─────────────────────────────────────────────────────────
def detect_dialect(sql):
    s = sql.upper()
    if re.search(r'DECLARE\s+@', s):            return "T-SQL (SQL Server)"
    if re.search(r'LANGUAGE\s+PLPGSQL', s):     return "PostgreSQL"
    if re.search(r'CREATE\s+OR\s+REPLACE', s):
        return "PostgreSQL" if "LANGUAGE" in s else "Oracle PL/SQL"
    if re.search(r'`\w+`', sql):                return "MySQL"
    if re.search(r'\$\$', sql):                 return "PostgreSQL"
    return "Auto-detected"


# ── SP Splitter ───────────────────────────────────────────────────────────────
def split_procedures(sql):
    pat = re.compile(
        r'(CREATE\s+(?:OR\s+REPLACE\s+)?(?:PROCEDURE|PROC|FUNCTION)\s+([\w\.\[\]"` ]+))',
        re.IGNORECASE
    )
    matches = list(pat.finditer(sql))
    if not matches:
        return [("UnnamedProcedure", sql)]
    procs = []
    for i, m in enumerate(matches):
        start = m.start()
        end   = matches[i+1].start() if i+1 < len(matches) else len(sql)
        name  = strip_name(m.group(2)).strip()
        procs.append((name, sql[start:end]))
    return procs


# ── Extraction ────────────────────────────────────────────────────────────────
SKIP_WORDS = {
    'SELECT','WHERE','SET','ON','AND','OR','NOT','IN','AS','WITH','BY','HAVING',
    'UNION','ALL','DISTINCT','TOP','NULL','BEGIN','END','IF','ELSE','THEN','CASE',
    'WHEN','RETURN','DECLARE','PRINT','GO','USE','OUTPUT','DEFAULT','VALUES',
    'EXISTS','COALESCE','ISNULL','CAST','CONVERT','GETDATE','SYSDATETIME','NEWID',
    'NOLOCK','READPAST','UPDLOCK','ROWLOCK','TABLOCK','INTO','FROM','EXEC',
    'EXECUTE','PROCEDURE','PROC','FUNCTION','TRIGGER','VIEW','TABLE','INDEX',
    'DATABASE','SCHEMA','SCOPE_IDENTITY','OBJECT_ID','ISNUMERIC','ISDATE','LEN',
    'LTRIM','RTRIM','UPPER','LOWER','SUBSTRING','CHARINDEX','REPLACE','STUFF',
    'DATEDIFF','DATEADD','YEAR','MONTH','DAY','COUNT','SUM','AVG','MIN','MAX',
    'ROW_NUMBER','RANK','DENSE_RANK','NTILE','LAG','LEAD','OVER','PARTITION',
    'INSERTED','DELETED','IDENTITY','CONSTRAINT','PRIMARY','FOREIGN','KEY',
    'NVARCHAR','VARCHAR','INT','BIGINT','DATETIME','DATE','BIT','DECIMAL','FLOAT',
    'CHAR','TEXT','UNIQUEIDENTIFIER','MONEY','SMALLINT','TINYINT','IMAGE',
    'OBJECT_NAME','DB_NAME','USER_NAME','SUSER_NAME','HOST_NAME'
}

OP_PATTERNS = [
    (r'\bFROM\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',           'SELECT'),
    (r'\bINNER\s+JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',   'SELECT'),
    (r'\bLEFT\s+(?:OUTER\s+)?JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',   'SELECT'),
    (r'\bRIGHT\s+(?:OUTER\s+)?JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',  'SELECT'),
    (r'\bFULL\s+(?:OUTER\s+)?JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',   'SELECT'),
    (r'\bCROSS\s+JOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',   'SELECT'),
    (r'\bJOIN\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',            'SELECT'),
    (r'\bINSERT\s+INTO\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',  'INSERT'),
    (r'\bUPDATE\s+((?:[\w\[\]]+\.)*[\w\[\]]+)\s+SET\b', 'UPDATE'),
    (r'\bDELETE\s+FROM\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',  'DELETE'),
    (r'\bMERGE\s+(?:INTO\s+)?((?:[\w\[\]]+\.)*[\w\[\]]+)', 'MERGE'),
    (r'\bTRUNCATE\s+TABLE\s+((?:[\w\[\]]+\.)*[\w\[\]]+)', 'TRUNCATE'),
    (r'\bUSING\s+((?:[\w\[\]]+\.)*[\w\[\]]+)',           'SELECT'),
]

TEMP_OP_PATTERNS = [
    (r'\bFROM\s+(#[\w]+)',                              'SELECT'),
    (r'\bJOIN\s+(#[\w]+)',                              'SELECT'),
    (r'\bINSERT\s+INTO\s+(#[\w]+)',                     'INSERT'),
    (r'\bSELECT\b.*?\bINTO\s+(#[\w]+)',                 'INSERT'),
    (r'\bUPDATE\s+(#[\w]+)\s+SET\b',                   'UPDATE'),
    (r'\bDELETE\s+FROM\s+(#[\w]+)',                     'DELETE'),
    (r'\bCREATE\s+TABLE\s+(#[\w]+)',                    'CREATE'),
    (r'\bDROP\s+TABLE\s+(?:IF\s+EXISTS\s+)?(#[\w]+)',  'DROP'),
    (r'\bTRUNCATE\s+TABLE\s+(#[\w]+)',                  'TRUNCATE'),
]


def clean_sql(sql):
    sql = re.sub(r'/\*.*?\*/', ' ', sql, flags=re.DOTALL)
    sql = re.sub(r'--[^\n]*', ' ', sql)
    sql = re.sub(r'\s+', ' ', sql)
    return sql.strip()


def parse_table_ref(raw):
    """Returns (schema, table_base, full_name) from a raw match."""
    raw = strip_name(raw).strip()
    parts = [p for p in raw.split('.') if p]
    if len(parts) >= 3:
        schema = parts[-2]
        base   = parts[-1]
        full   = f"{schema}.{base}"
    elif len(parts) == 2:
        schema = parts[0]
        base   = parts[1]
        full   = f"{schema}.{base}"
    else:
        schema = ""
        base   = parts[0] if parts else raw
        full   = base
    return schema.upper(), base.upper(), full


def extract_all(sql):
    clean = clean_sql(sql)

    # ── CTE names
    cte_names = set()
    for m in re.finditer(r'\bWITH\b\s+([\w\[\]]+)\s+AS\s*\(', clean, re.IGNORECASE):
        cte_names.add(strip_name(m.group(1)).upper())
    for m in re.finditer(r',\s*([\w\[\]]+)\s+AS\s*\(', clean, re.IGNORECASE):
        cte_names.add(strip_name(m.group(1)).upper())

    # physical: key=full_name_upper → {schema, base, ops, aliases, columns}
    physical    = {}
    temp_warned = {}   # #tempname → set of ops (for warning notes only)
    alias_map   = {}   # alias_upper → full_name_upper in physical

    # ── Physical table extraction
    for pat, op in OP_PATTERNS:
        for m in re.finditer(pat, clean, re.IGNORECASE):
            raw = m.group(1).strip()
            schema, base, full = parse_table_ref(raw)
            if base in SKIP_WORDS or base in cte_names or is_var(base) or is_temp(base) or len(base) < 2:
                continue
            key = full.upper()
            if key not in physical:
                physical[key] = {"schema": schema, "base": base, "full": full,
                                  "ops": set(), "aliases": set(), "columns": set()}
            physical[key]["ops"].add(op)

    # ── Temp table ops (for warning sheet only)
    for pat, op in TEMP_OP_PATTERNS:
        for m in re.finditer(pat, clean, re.IGNORECASE | re.DOTALL):
            tname = m.group(1).upper()
            if tname not in temp_warned:
                temp_warned[tname] = set()
            temp_warned[tname].add(op)

    # ── Alias extraction
    alias_pat = re.compile(
        r'(?:FROM|JOIN|UPDATE|MERGE)\s+((?:[\w\[\]]+\.)*[\w\[\]]+)\s+(?:AS\s+)?([\w]+)'
        r'(?=\s|\(|$)',
        re.IGNORECASE
    )
    for m in alias_pat.finditer(clean):
        traw  = strip_name(m.group(1))
        alias = strip_name(m.group(2)).upper()
        if alias in SKIP_WORDS or is_temp(traw):
            continue
        _, _, full = parse_table_ref(traw)
        key = full.upper()
        if key in physical and alias != physical[key]["base"].upper():
            physical[key]["aliases"].add(alias)
            alias_map[alias] = key

    # ── Column extraction — qualified references (prefix.col)
    for m in re.finditer(r'\b([\w\[\]]+)\.([\w\[\]]+)\b', clean):
        prefix = strip_name(m.group(1)).upper()
        col    = strip_name(m.group(2)).upper()
        if col in SKIP_WORDS or is_var(col) or not re.match(r'^[A-Z_][A-Z0-9_]*$', col):
            continue
        # resolve via alias map first, then direct table name match
        target = alias_map.get(prefix)
        if not target:
            for k in physical:
                if k == prefix or k.endswith('.' + prefix):
                    target = k
                    break
        if target and target in physical:
            physical[target]["columns"].add(col)

    # ── Column extraction — unqualified (SELECT list, SET, INSERT col list)
    unresolved_cols = set()

    for m in re.finditer(r'\bSELECT\b(.*?)\bFROM\b', clean, re.IGNORECASE | re.DOTALL):
        block = re.sub(r'\(.*?\)', '', m.group(1), flags=re.DOTALL)
        for token in re.split(r'[,\s]+', block):
            token = strip_name(token.split('.')[-1]).upper()
            if token and token not in SKIP_WORDS and token != '*' and re.match(r'^[A-Z_][A-Z0-9_]*$', token):
                unresolved_cols.add(token)

    for m in re.finditer(r'\bSET\s+([\w\[\]]+)\s*=', clean, re.IGNORECASE):
        col = strip_name(m.group(1)).upper()
        if col not in SKIP_WORDS:
            unresolved_cols.add(col)

    for m in re.finditer(r'\bINSERT\s+INTO\s+[\w\.\[\]]+\s*\((.*?)\)', clean, re.IGNORECASE | re.DOTALL):
        for c in m.group(1).split(','):
            c = strip_name(c).upper().strip()
            if c and re.match(r'^[A-Z_][A-Z0-9_]*$', c) and c not in SKIP_WORDS:
                unresolved_cols.add(c)

    # Remove already-mapped cols and skip words
    mapped = set(col for info in physical.values() for col in info["columns"])
    unresolved_cols -= mapped
    unresolved_cols -= SKIP_WORDS

    # Dynamic SQL flag
    dynamic = bool(re.search(r'EXEC\s*\(|EXECUTE\s*\(|sp_executesql', clean, re.IGNORECASE))

    return physical, temp_warned, cte_names, unresolved_cols, dynamic


# ── Excel Builder ─────────────────────────────────────────────────────────────
def build_excel(all_results, output_path):
    wb = Workbook()

    # ════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Stored Procedure Analyzer v3 — Extraction Report"
    c.font  = Font(bold=True, name="Arial", size=14, color=C_HEADER_FG)
    c.fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    |    Procedures: {len(all_results)}"
    c.font  = Font(italic=True, name="Arial", size=9, color="595959")
    c.alignment = Alignment(horizontal="center")

    hdrs = ["#","Procedure Name","Source","Dialect","Physical Tables","CTEs",
            "Temp Tables (⚠ note only)","Dynamic SQL?"]
    for ci, h in enumerate(hdrs, 1):
        style_header(ws.cell(row=4, column=ci, value=h))
    ws.row_dimensions[4].height = 22

    for ri, (pname, source, dialect, phys, tmp, ctes, unres, dynamic) in enumerate(all_results, 1):
        vals = [ri, pname,
                Path(source).name if source != "stdin" else "pasted input",
                dialect, len(phys), len(ctes), len(tmp),
                "⚠ YES" if dynamic else "No"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri+4, column=ci, value=v)
            style_data(c, ri)
        # temp count in orange
        ws.cell(row=ri+4, column=7).font = Font(
            bold=bool(tmp), color=C_WARN_FG if tmp else "404040", name="Arial", size=10)
        if dynamic:
            ws.cell(row=ri+4, column=8).font = Font(bold=True, color="C00000", name="Arial", size=10)

    for i, w in enumerate([5,35,22,20,16,10,22,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet 2 — Physical Tables & Operations
    # ════════════════════════════════════════════════════════
    wt = wb.create_sheet("Physical Tables")
    wt.sheet_view.showGridLines = False
    hdrs = ["Procedure","Schema","Table Name","Full Reference","Operation(s)","Aliases Used"]
    for ci, h in enumerate(hdrs, 1):
        style_header(wt.cell(row=1, column=ci, value=h))

    row = 2
    for pname, source, dialect, phys, tmp, ctes, unres, dynamic in all_results:
        for key, info in sorted(phys.items()):
            ops     = ', '.join(sorted(info["ops"]))
            aliases = ', '.join(sorted(info["aliases"])) or '—'
            full_ref = f"{info['schema']}.{info['base']}" if info['schema'] else info['base']
            vals = [pname, info['schema'] or '(none)', info['base'], full_ref, ops, aliases]
            for ci, v in enumerate(vals, 1):
                c = wt.cell(row=row, column=ci, value=v)
                style_data(c, row)
            wt.cell(row=row, column=5).font = Font(
                bold=True, color=op_color(info["ops"]), name="Arial", size=10)
            row += 1

    for i, w in enumerate([32,14,26,28,26,20], 1):
        wt.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet 3 — Columns Detail
    # ════════════════════════════════════════════════════════
    wc = wb.create_sheet("Columns Detail")
    wc.sheet_view.showGridLines = False
    hdrs = ["Procedure","Schema","Table Name","Column Name","Mapped Via","Operation(s)"]
    for ci, h in enumerate(hdrs, 1):
        style_header(wc.cell(row=1, column=ci, value=h))

    row = 2
    for pname, source, dialect, phys, tmp, ctes, unres, dynamic in all_results:
        for key, info in sorted(phys.items()):
            cols = sorted(info["columns"])
            schema = info['schema'] or '(none)'
            base   = info['base']
            ops    = ', '.join(sorted(info["ops"])) or '—'

            if not cols:
                vals = [pname, schema, base, '(no columns resolved)', 'alias.col / table.col', ops]
                for ci, v in enumerate(vals, 1):
                    c = wc.cell(row=row, column=ci, value=v)
                    style_data(c, row)
                    if ci == 4:
                        c.font = Font(italic=True, color="808080", name="Arial", size=10)
                row += 1
            else:
                for col in cols:
                    vals = [pname, schema, base, col, 'alias.col / table.col', ops]
                    for ci, v in enumerate(vals, 1):
                        c = wc.cell(row=row, column=ci, value=v)
                        style_data(c, row)
                    row += 1

        # Unresolved columns
        if unres:
            for col in sorted(unres):
                vals = [pname, '—', '(table not determined)', col, 'unqualified token', '—']
                for ci, v in enumerate(vals, 1):
                    c = wc.cell(row=row, column=ci, value=v)
                    style_data(c, row)
                    if ci in [2, 3, 5, 6]:
                        c.font = Font(italic=True, color="808080", name="Arial", size=10)
                row += 1

    for i, w in enumerate([32,14,26,28,22,22], 1):
        wc.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet 4 — Schema Breakdown (Schema → Tables → Columns)
    # ════════════════════════════════════════════════════════
    wb_schema = wb.create_sheet("Schema Breakdown")
    wb_schema.sheet_view.showGridLines = False

    wb_schema.merge_cells("A1:E1")
    c = wb_schema["A1"]
    c.value = "Schema Breakdown — Physical Tables & Columns by Schema"
    c.font  = Font(bold=True, name="Arial", size=13, color=C_HEADER_FG)
    c.fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    wb_schema.row_dimensions[1].height = 30

    hdrs = ["Schema","Table Name","Column Name","Operations on Table","Found In Procedure(s)"]
    for ci, h in enumerate(hdrs, 1):
        style_header(wb_schema.cell(row=2, column=ci, value=h))

    # Build schema → table → {columns, ops, procs}
    schema_map = defaultdict(lambda: defaultdict(lambda: {"columns": set(), "ops": set(), "procs": set()}))
    for pname, source, dialect, phys, tmp, ctes, unres, dynamic in all_results:
        for key, info in phys.items():
            schema = info['schema'] or '(no schema)'
            base   = info['base']
            schema_map[schema][base]["columns"].update(info["columns"])
            schema_map[schema][base]["ops"].update(info["ops"])
            schema_map[schema][base]["procs"].add(pname)

    row = 3
    for schema in sorted(schema_map.keys()):
        # Schema group header row
        wb_schema.merge_cells(f"A{row}:E{row}")
        c = wb_schema.cell(row=row, column=1, value=f"  SCHEMA:  {schema}")
        style_schema_group(c)
        # Style the merged header row background via the first cell only
        c.fill   = PatternFill("solid", fgColor=C_SCHEMA_BG)
        c.border = thin_border()
        wb_schema.row_dimensions[row].height = 20
        row += 1

        for tname in sorted(schema_map[schema].keys()):
            tinfo   = schema_map[schema][tname]
            cols    = sorted(tinfo["columns"])
            ops     = ', '.join(sorted(tinfo["ops"]))
            procs   = ', '.join(sorted(tinfo["procs"]))

            if not cols:
                vals = [schema, tname, '(no columns resolved)', ops, procs]
                for ci, v in enumerate(vals, 1):
                    c = wb_schema.cell(row=row, column=ci, value=v)
                    style_data(c, row)
                    if ci == 3:
                        c.font = Font(italic=True, color="808080", name="Arial", size=10)
                wb_schema.cell(row=row, column=4).font = Font(
                    bold=True, color=op_color(tinfo["ops"]), name="Arial", size=10)
                row += 1
            else:
                for i, col in enumerate(cols):
                    # Only repeat schema+table on first column row
                    s_val = schema if i == 0 else ""
                    t_val = tname  if i == 0 else ""
                    o_val = ops    if i == 0 else ""
                    p_val = procs  if i == 0 else ""
                    vals = [s_val, t_val, col, o_val, p_val]
                    for ci, v in enumerate(vals, 1):
                        c = wb_schema.cell(row=row, column=ci, value=v)
                        style_data(c, row)
                    if i == 0:
                        wb_schema.cell(row=row, column=4).font = Font(
                            bold=True, color=op_color(tinfo["ops"]), name="Arial", size=10)
                    row += 1

    for i, w in enumerate([16,26,28,26,40], 1):
        wb_schema.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet 5 — Temp Table Warnings (note only, not real tables)
    # ════════════════════════════════════════════════════════
    ww = wb.create_sheet("⚠ Temp Table Notes")
    ww.sheet_view.showGridLines = False

    ww.merge_cells("A1:D1")
    c = ww["A1"]
    c.value = "⚠  Temporary Tables Detected — For Reference Only (NOT physical schema objects)"
    c.font  = Font(bold=True, name="Arial", size=11, color=C_WARN_FG)
    c.fill  = PatternFill("solid", fgColor=C_WARN_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ww.row_dimensions[1].height = 24

    ww.merge_cells("A2:D2")
    c = ww["A2"]
    c.value = ("Temp tables exist only during stored procedure execution. "
               "They are excluded from Physical Tables, Columns Detail, and Schema Breakdown. "
               "Review manually for data flow / lineage purposes.")
    c.font  = Font(italic=True, name="Arial", size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ww.row_dimensions[2].height = 30

    hdrs = ["Procedure","Temp Table Name","Operations Detected","Analyst Note"]
    for ci, h in enumerate(hdrs, 1):
        style_header(ww.cell(row=3, column=ci, value=h), bg=C_WARN_FG, fg=C_HEADER_FG)

    row = 4
    any_temp = False
    for pname, source, dialect, phys, tmp, ctes, unres, dynamic in all_results:
        for tname, ops in sorted(tmp.items()):
            any_temp = True
            op_str = ', '.join(sorted(ops))
            note = ("Intermediate staging — data flows through this table" 
                    if "INSERT" in ops and "SELECT" in ops
                    else "Created and used within SP" if "CREATE" in ops
                    else "Referenced but not created here" if "SELECT" in ops
                    else "Written to within SP")
            vals = [pname, tname, op_str, note]
            for ci, v in enumerate(vals, 1):
                c = ww.cell(row=row, column=ci, value=v)
                style_warn(c)
            row += 1

    if not any_temp:
        c = ww.cell(row=4, column=1, value="No temporary tables (#table) detected across all procedures.")
        c.font = Font(italic=True, color="808080", name="Arial", size=10)

    for i, w in enumerate([32,24,26,50], 1):
        ww.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet 6 — CTEs
    # ════════════════════════════════════════════════════════
    wq = wb.create_sheet("CTEs")
    wq.sheet_view.showGridLines = False
    for ci, h in enumerate(["Procedure","CTE Name","Note"], 1):
        style_header(wq.cell(row=1, column=ci, value=h))
    row = 2
    any_cte = False
    for pname, source, dialect, phys, tmp, ctes, unres, dynamic in all_results:
        for cte in sorted(ctes):
            any_cte = True
            for ci, v in enumerate([pname, cte, "Virtual — not a physical table or temp table"], 1):
                style_data(wq.cell(row=row, column=ci, value=v), row)
            row += 1
    if not any_cte:
        wq["A2"] = "No CTEs detected"
        wq["A2"].font = Font(italic=True, color="808080")
    for i, w in enumerate([32,28,42], 1):
        wq.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════
    # Sheet 7 — Legend
    # ════════════════════════════════════════════════════════
    wl = wb.create_sheet("Legend")
    wl.sheet_view.showGridLines = False
    wl.merge_cells("A1:D1")
    c = wl["A1"]
    c.value = "Operation Color Legend & Analyst Notes"
    style_header(c, size=12)
    wl.row_dimensions[1].height = 26

    items = [
        ("SELECT",   OP_COLORS["SELECT"],   "Data read — FROM, JOIN, USING"),
        ("INSERT",   OP_COLORS["INSERT"],   "Data write — INSERT INTO"),
        ("UPDATE",   OP_COLORS["UPDATE"],   "Data modification — UPDATE ... SET"),
        ("DELETE",   OP_COLORS["DELETE"],   "Data removal — DELETE FROM"),
        ("MERGE",    OP_COLORS["MERGE"],    "Upsert — MERGE INTO ... USING"),
        ("TRUNCATE", OP_COLORS["TRUNCATE"], "Full wipe — TRUNCATE TABLE"),
        ("CREATE",   OP_COLORS["CREATE"],   "Object creation — CREATE TABLE"),
        ("DROP",     OP_COLORS["DROP"],     "Object removal — DROP TABLE"),
    ]
    notes = [
        ("",""),
        ("Physical Tables",  "Permanent schema objects only — temp tables excluded from these sheets"),
        ("Schema Breakdown", "Groups physical tables by schema; shows all columns per table across all SPs"),
        ("⚠ Temp Table Notes","Temp (#) tables shown here as reference only — not counted as real tables"),
        ("Unresolved Cols",  "Column found in SQL but table could not be determined from context"),
        ("Dynamic SQL ⚠",   "EXEC() / sp_executesql — contents cannot be statically analyzed"),
        ("CTE",              "Common Table Expression — virtual query alias, not a stored object"),
    ]
    for ri, (op, color, desc) in enumerate(items, 2):
        wl.cell(row=ri, column=1, value=op).font  = Font(bold=True, color=color, name="Arial", size=11)
        wl.cell(row=ri, column=1).alignment        = Alignment(horizontal="center")
        wl.cell(row=ri, column=2, value="●").font  = Font(color=color, size=14)
        wl.cell(row=ri, column=3, value=desc).font = Font(name="Arial", size=10)

    for ri, (label, note) in enumerate(notes, len(items)+3):
        if label:
            wl.cell(row=ri, column=1, value=label).font = Font(bold=True, color=C_SUBHDR_BG, name="Arial", size=10)
        wl.cell(row=ri, column=3, value=note).font = Font(italic=True, name="Arial", size=10, color="404040")

    for i, w in enumerate([18,5,68], 1):
        wl.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n✅  Excel report saved → {output_path}\n")


# ── Process & Main ────────────────────────────────────────────────────────────
def process_sql(sql, source, forced_dialect):
    dialect = detect_dialect(sql)
    if forced_dialect:
        dialect = forced_dialect
    procs   = split_procedures(sql)
    results = []
    for pname, body in procs:
        phys, tmp, ctes, unres, dynamic = extract_all(body)
        src_label = Path(source).name if source != "stdin" else "pasted"
        print(f"  [{src_label}] → {pname}: "
              f"{len(phys)} physical table(s), {len(tmp)} temp (note only), "
              f"{len(ctes)} CTE(s), dialect={dialect}"
              + (" ⚠ dynamic SQL" if dynamic else ""))
        results.append((pname, source, dialect, phys, tmp, ctes, unres, dynamic))
    return results


def main():
    parser = argparse.ArgumentParser(description="SP Analyzer v3 → Excel")
    parser.add_argument("files",     nargs="*",  help=".sql files to process")
    parser.add_argument("--dialect", default="", help="Force dialect label")
    parser.add_argument("--output",  default="sp_analysis_v3.xlsx")
    args = parser.parse_args()

    all_results = []
    if args.files:
        for fp in args.files:
            sql = Path(fp).read_text(encoding="utf-8")
            all_results.extend(process_sql(sql, fp, args.dialect))
    else:
        print("Paste your stored procedure(s) below.")
        print("Press Ctrl+D (Mac/Linux) or Ctrl+Z+Enter (Windows) when done:\n")
        lines = []
        try:
            while True:
                lines.append(input())
        except EOFError:
            pass
        sql = "\n".join(lines)
        if not sql.strip():
            print("No SQL provided. Exiting.")
            sys.exit(1)
        all_results.extend(process_sql(sql, "stdin", args.dialect))

    if not all_results:
        print("Nothing to process.")
        sys.exit(1)

    build_excel(all_results, args.output)


if __name__ == "__main__":
    main()
